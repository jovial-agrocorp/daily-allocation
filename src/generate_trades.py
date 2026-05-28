import os
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from neon import Neon

SRC_DIR = os.path.dirname(os.path.abspath(__file__))
CUTOFF_HOUR   = 8  # Before 8am, treat as previous trading day
SF_ACCOUNTS   = ["TMG30982", "TMG30985", "TMG31991"]

# Most contracts come from Neon as a decimal (e.g. 6.2600 = 626.00 c/bu) so need *100.
# Add exceptions here for contracts already in their native display unit.
PRICE_MULTIPLIER = {
    "RS": 1,    # Canola: Neon returns CAD/tonne directly (e.g. 739.10)
}
_DEFAULT_PRICE_MULT = 100


def _price_mult(contract):
    code = contract[:-3]  # strip month letter + 2-digit year (e.g. "RSN26" -> "RS")
    return PRICE_MULTIPLIER.get(code, _DEFAULT_PRICE_MULT)

FUTURES_INFO_COLS = ["Account Number", "Trade Date", "Contract", "Commodity Name", "Long", "Short", "Price"]
OPTIONS_INFO_COLS = ["Account Number", "Trade Date", "Contract", "Commodity Name", "Long", "Short", "Price", "Strike", "Put/Call"]
EDITABLE_COLS     = ["Book", "Strategy", "New AGP", "New AGP Sub Contract", "New AGS", "New AGS Sub Contract"]
FUTURES_EDITABLE  = EDITABLE_COLS + ["Split Qty"]
OPTIONS_EDITABLE  = EDITABLE_COLS


def get_trading_date():
    now = datetime.now()
    if now.hour < CUTOFF_HOUR:
        return (now - timedelta(days=1)).strftime("%Y-%m-%d")
    return now.strftime("%Y-%m-%d")


def _fetch_aggregated_trades(trade_date=None):
    neon = Neon()
    neon.connect()
    trade_date = trade_date or get_trading_date()
    print(f"Using trade date: {trade_date}")
    result = neon.get_trades(trade_date)

    if not result["success"]:
        raise RuntimeError(result["error"])

    futures = [t for t in result["futures"] if t["account_number"] in SF_ACCOUNTS]
    options = [t for t in result["options"]  if t["account_number"] in SF_ACCOUNTS]

    futures_rows = []
    for t in futures:
        futures_rows.append({
            "Account Number":       t["account_number"],
            "Trade Date":           t["trade_date"],
            "Contract":             t["contract"],
            "Commodity Name":       t["commodity_name"],
            "Long":                 t["contracts"] if t["contracts"] > 0 else None,
            "Short":                abs(t["contracts"]) if t["contracts"] < 0 else None,
            "Price":                round(t["trade_price"] * _price_mult(t["contract"]), 2),
        })

    options_rows = []
    for t in options:
        options_rows.append({
            "Account Number":       t["account_number"],
            "Trade Date":           t["trade_date"],
            "Contract":             t["contract"],
            "Commodity Name":       t["commodity_name"],
            "Long":                 t["contracts"] if t["contracts"] > 0 else None,
            "Short":                abs(t["contracts"]) if t["contracts"] < 0 else None,
            "Price":                round(t["trade_price"] * _price_mult(t["contract"]), 2),
            "Strike":               t["strike_price"],
            "Put/Call":             t["call_or_put"],
        })

    def _sum_or_none(s):
        return s.sum() if s.notna().any() else None

    def _aggregate(rows, group_cols, info_cols, editable_cols):
        df = pd.DataFrame(rows)
        agg = df.groupby(group_cols, as_index=False).agg(
            Long=("Long", _sum_or_none),
            Short=("Short", _sum_or_none),
        )
        for col in editable_cols:
            agg[col] = None
        return agg[info_cols + editable_cols].sort_values("Contract").reset_index(drop=True)

    futures_group_cols = ["Account Number", "Trade Date", "Contract", "Commodity Name", "Price"]
    options_group_cols = ["Account Number", "Trade Date", "Contract", "Commodity Name", "Price", "Strike", "Put/Call"]

    df_futures = _aggregate(futures_rows, futures_group_cols, FUTURES_INFO_COLS, FUTURES_EDITABLE) if futures_rows else pd.DataFrame(columns=FUTURES_INFO_COLS + FUTURES_EDITABLE)
    df_options = _aggregate(options_rows, options_group_cols, OPTIONS_INFO_COLS,  OPTIONS_EDITABLE) if options_rows else pd.DataFrame(columns=OPTIONS_INFO_COLS  + OPTIONS_EDITABLE)

    return df_futures, df_options, trade_date, len(futures), len(options)


def _style_sheet(ws, df, info_cols, all_cols):
    gray_fill   = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)

    for col_idx, col_name in enumerate(all_cols, 1):
        col_letter = get_column_letter(col_idx)
        is_info    = col_name in info_cols

        for row_idx in range(1, len(df) + 2):
            cell = ws[f"{col_letter}{row_idx}"]
            cell.alignment = Alignment(horizontal="left")
            if is_info and row_idx > 1:
                cell.fill = gray_fill

        ws[f"{col_letter}1"].font = header_font

        max_len = max(
            len(str(col_name)),
            max((len(str(df.iloc[r][col_name])) for r in range(len(df))
                 if col_name in df.columns and pd.notna(df.iloc[r][col_name])), default=0)
        )
        ws.column_dimensions[col_letter].width = max_len + 4

    ws.auto_filter.ref = ws.dimensions


def generate_trades(trade_date=None, output_path=None):
    try:
        df_futures, df_options, trade_date, n_futures, n_options = _fetch_aggregated_trades(trade_date)
    except RuntimeError as e:
        print(f"ERROR: Failed to fetch trades: {e}")
        return

    save_path = output_path or os.path.join(SRC_DIR, f"neon_trades_{trade_date}.xlsx")
    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
        df_futures.to_excel(writer, index=False, sheet_name="Futures")
        df_options.to_excel(writer, index=False, sheet_name="Options")
        _style_sheet(writer.sheets["Futures"], df_futures, FUTURES_INFO_COLS, FUTURES_INFO_COLS + FUTURES_EDITABLE)
        _style_sheet(writer.sheets["Options"], df_options, OPTIONS_INFO_COLS,  OPTIONS_INFO_COLS + OPTIONS_EDITABLE)

    print(f"{n_futures} futures, {n_options} options fetched for {', '.join(SF_ACCOUNTS)}")
    return save_path


def format_trades_text(trade_date=None):
    df_futures, df_options, trade_date, _, _ = _fetch_aggregated_trades(trade_date)

    lines = []

    if not df_futures.empty:
        lines.append(f"FUTURES  {trade_date}")
        lines.append("")
        for acct, grp in df_futures.groupby("Account Number"):
            lines.append(f"[{acct}]")
            lines.append(f"{'Contract':<10} {'Side':<5} {'Qty':>6}  {'Price':>9}")
            lines.append("-" * 36)
            for _, row in grp.iterrows():
                long_qty  = int(row["Long"])  if pd.notna(row["Long"])  else None
                short_qty = int(row["Short"]) if pd.notna(row["Short"]) else None
                if long_qty and short_qty:
                    lines.append(f"{row['Contract']:<10} {'Long':<5} {long_qty:>6}  {row['Price']:>9}")
                    lines.append(f"{'':<10} {'Short':<5} {short_qty:>6}")
                elif long_qty:
                    lines.append(f"{row['Contract']:<10} {'Long':<5} {long_qty:>6}  {row['Price']:>9}")
                else:
                    lines.append(f"{row['Contract']:<10} {'Short':<5} {short_qty:>6}  {row['Price']:>9}")
            lines.append("")

    if not df_options.empty:
        lines.append(f"OPTIONS  {trade_date}")
        lines.append("")
        for acct, grp in df_options.groupby("Account Number"):
            lines.append(f"[{acct}]")
            lines.append(f"{'Contract':<10} {'Side':<5} {'Qty':>6}  {'Price':>9}  {'Strike':>9}  P/C")
            lines.append("-" * 52)
            for _, row in grp.iterrows():
                long_qty  = int(row["Long"])  if pd.notna(row["Long"])  else None
                short_qty = int(row["Short"]) if pd.notna(row["Short"]) else None
                qty  = long_qty or short_qty
                side = "Long" if long_qty else "Short"
                lines.append(f"{row['Contract']:<10} {side:<5} {qty:>6}  {row['Price']:>9}  {row['Strike']:>9}  {row['Put/Call']}")
            lines.append("")

    if not lines:
        return [f"No trades found for {trade_date}."]

    # Chunk into messages that fit Telegram's 4096-char limit
    LIMIT = 3800
    messages = []
    chunk = []
    chunk_len = 0
    for line in lines:
        n = len(line) + 1
        if chunk_len + n > LIMIT and chunk:
            messages.append("```\n" + "\n".join(chunk) + "\n```")
            chunk = []
            chunk_len = 0
        chunk.append(line)
        chunk_len += n
    if chunk:
        messages.append("```\n" + "\n".join(chunk) + "\n```")

    return messages


if __name__ == "__main__":
    generate_trades()
